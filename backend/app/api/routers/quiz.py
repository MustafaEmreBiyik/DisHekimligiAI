"""
Quiz Router
===========
Serves MCQ and Open-Ended questions. S8B implementation.
Answer keys and protected fields are strictly omitted.
"""

from fastapi import APIRouter, status, Depends, HTTPException, UploadFile
from pydantic import BaseModel, Field
from typing import Dict, List, Optional
import json
import logging
from pathlib import Path
from datetime import datetime
import re
import unicodedata
from collections import defaultdict
from sqlalchemy.orm import Session

from app.api.deps import get_current_user_context, AuthenticatedUser, get_db, require_roles
from app.services.composite_score_service import calculate_composite_score, CompositeScoreResult
from app.services.topic_accuracy_service import get_topic_accuracy, TopicAccuracyResult
from app.services.question_case_mapping_service import (
    get_question_case_mappings,
    create_mapping,
    delete_mapping,
    MappingQueryResult,
    QuestionNotFoundError,
    DuplicateMappingError,
    MappingNotFoundError,
)
from db.database import ExamResult, UserRole, User, Question, QuestionType, QuizAttempt, QuizAnswer, GradingStatus, AIScoringLog, Notification, RubricVersion, ReviewSchedule

logger = logging.getLogger(__name__)

router = APIRouter()

QUESTIONS_FILE = Path(__file__).resolve().parents[3] / "data" / "question_bank" / "mcq_questions.json"

from app.constants import TOPIC_LABELS as TOPIC_MAP


# ── Pydantic schemas ──────────────────────────────────────────────────────────

class StudentQuestion(BaseModel):
    """Student-safe schema: no answer keys or protected fields."""
    id: str
    topic: str
    question: str
    options: List[str]
    question_type: str = "MCQ"
    difficulty: str = "medium"
    bloom_level: str = "apply"


class SubmitRequest(BaseModel):
    answers: Dict[str, str]   # {question_id: selected_option or free_text}


class QuestionFeedback(BaseModel):
    id: str
    topic: str
    question: str
    question_type: str = "MCQ"
    selected_option: Optional[str] = None
    is_correct: Optional[bool] = None
    feedback: Optional[str] = None
    grading_status: str = "PUBLISHED"
    instructor_score: Optional[int] = None
    instructor_feedback: Optional[str] = None
    answer_id: Optional[int] = None  # S10-B: needed for explanation lookup


class SubmitResponse(BaseModel):
    attempt_id: int
    score: Optional[int] = None
    total: int
    percentage: Optional[int] = None
    overall_status: str = "PUBLISHED"
    results: List[QuestionFeedback]


class GradingQueueItem(BaseModel):
    answer_id: int
    attempt_id: int
    question_id: str
    question_text: str
    student_response: str
    rubric_guide: Optional[str] = None
    model_answer_outline: Optional[str] = None
    max_score: int
    submitted_at: Optional[str] = None


class GradeSubmission(BaseModel):
    instructor_score: int
    instructor_feedback: str
    publish: bool


class AttemptSummary(BaseModel):
    attempted: bool
    last_score: Optional[float] = None
    attempt_count: int


class QuestionBankEntry(BaseModel):
    """Student-safe question with attempt context. No answer keys exposed."""
    question_id: str
    question_text: str
    question_type: str
    topic_id: str
    bloom_level: str
    difficulty: str
    max_score: int
    options_json: Optional[List[str]] = None
    attempt_summary: AttemptSummary


class InstructorQuestionCreateRequest(BaseModel):
    question_type: str = QuestionType.OPEN_ENDED.value
    question_id: Optional[str] = None
    question_text: str
    topic_id: str
    competency_areas: List[str] = Field(default_factory=list)
    bloom_level: str
    difficulty: str
    safety_category: str
    unit_id: Optional[str] = None      # T-2A: Sprint-1 tagging — ünite
    week_number: Optional[int] = None  # T-2A: Sprint-1 tagging — hafta
    rubric_guide: Optional[str] = None
    model_answer_outline: Optional[str] = None
    instructor_explanation: Optional[str] = None
    options: List[str] = Field(default_factory=list)
    correct_option: Optional[str] = None
    max_score: int = 10
    is_active: bool = True


class InstructorQuestionSummary(BaseModel):
    id: int                          # T-4B: DB primary key
    question_id: str
    question_type: str
    question_text: str
    topic_id: str
    competency_areas: List[str] = Field(default_factory=list)
    bloom_level: str
    difficulty: str
    safety_category: str
    unit_id: Optional[str] = None      # T-2A
    week_number: Optional[int] = None  # T-2A
    rubric_guide: Optional[str] = None
    model_answer_outline: Optional[str] = None
    instructor_explanation: Optional[str] = None
    options: List[str] = Field(default_factory=list)
    correct_option: Optional[str] = None
    max_score: int
    is_active: bool
    current_rubric_version: Optional[int] = None  # T-4B
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


def _forbidden() -> HTTPException:
    return HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")


def _safe_feedback(*, is_correct: bool) -> str:
    if is_correct:
        return "Dogru cevap. Benzer vakalarda ayni klinik yaklasimi surdurun."
    return "Bu yanit dogru degil. Konuyu yeniden gozden gecirip tekrar deneyin."


def _normalize_text(value: Optional[str]) -> str:
    return (value or "").strip()


def _normalize_str_list(values: Optional[List[str]]) -> List[str]:
    normalized: List[str] = []
    for value in values or []:
        cleaned = _normalize_text(value)
        if cleaned and cleaned not in normalized:
            normalized.append(cleaned)
    return normalized


def _slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_value.lower()).strip("-")
    return slug or "question"


def _question_to_summary(question: Question) -> InstructorQuestionSummary:
    return InstructorQuestionSummary(
        id=question.id,                        # T-4B
        question_id=question.question_id,
        question_type=question.question_type.value,
        question_text=question.question_text,
        topic_id=question.topic_id,
        competency_areas=question.competency_areas or [],
        bloom_level=question.bloom_level,
        difficulty=question.difficulty,
        safety_category=question.safety_category,
        unit_id=getattr(question, "unit_id", None),      # T-2A
        week_number=getattr(question, "week_number", None),  # T-2A
        rubric_guide=question.rubric_guide,
        model_answer_outline=question.model_answer_outline,
        instructor_explanation=question.instructor_explanation,
        options=question.options_json or [],
        correct_option=question.correct_option,
        max_score=question.max_score,
        is_active=question.is_active,
        current_rubric_version=getattr(question, 'current_rubric_version', None),  # T-4B
        created_at=question.created_at.isoformat() if question.created_at else None,
        updated_at=question.updated_at.isoformat() if question.updated_at else None,
    )


def _ensure_review_schedule(db: Session, user_id: str, question: Question) -> None:
    """Create a ReviewSchedule entry for this question if one doesn't already exist. S10-C."""
    existing = db.query(ReviewSchedule).filter(
        ReviewSchedule.user_id == user_id,
        ReviewSchedule.question_id == question.id,
    ).first()
    if not existing:
        db.add(ReviewSchedule(
            user_id=user_id,
            question_id=question.id,
            due_date=datetime.utcnow(),
            interval_days=1,
            ease_factor=2.5,
            repetitions=0,
            created_at=datetime.utcnow(),
        ))


def _generate_question_id(db: Session, question_type: QuestionType, topic_id: str, question_text: str) -> str:
    topic_slug = _slugify(topic_id)[:24]
    question_slug = _slugify(question_text)[:36]
    type_prefix = "oe" if question_type == QuestionType.OPEN_ENDED else "mcq"
    base_id = f"{type_prefix}-{topic_slug}-{question_slug}".strip("-")

    candidate = base_id
    suffix = 2
    while db.query(Question).filter(Question.question_id == candidate).first():
        candidate = f"{base_id}-{suffix}"
        suffix += 1

    return candidate


def _normalize_question_type(value: Optional[str]) -> QuestionType:
    normalized = _normalize_text(value).upper()
    if normalized == QuestionType.MCQ.value:
        return QuestionType.MCQ
    if normalized == QuestionType.OPEN_ENDED.value:
        return QuestionType.OPEN_ENDED
    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail="Question type must be MCQ or OPEN_ENDED",
    )


# ── Internal helpers ──────────────────────────────────────────────────────────

def _load_full_questions() -> List[dict]:
    """Load legacy full question data."""
    if not QUESTIONS_FILE.exists():
        logger.warning(f"Questions file not found: {QUESTIONS_FILE}")
        return []

    try:
        with open(QUESTIONS_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)

        questions = []
        for category_key, items in raw.items():
            topic_label = TOPIC_MAP.get(category_key, category_key.replace("_", " ").title())
            for item in items:
                questions.append({
                    "id": item.get("id", ""),
                    "topic": topic_label,
                    "question": item.get("question", ""),
                    "options": item.get("options", []),
                    "correct_option": item.get("correct_option", ""),
                    "explanation": item.get("explanation", ""),
                })
        return questions
    except Exception as e:
        logger.error(f"Failed to load questions: {e}")
        return []


# ── Student Endpoints ─────────────────────────────────────────────────────────

@router.get("/questions", response_model=List[StudentQuestion], status_code=status.HTTP_200_OK)
def get_questions(
    topic: Optional[str] = None,
    difficulty: Optional[str] = None,
    question_type: Optional[str] = None,
    bloom_level: Optional[str] = None,
    search: Optional[str] = None,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)),
    db: Session = Depends(get_db),
):
    db_questions = db.query(Question).filter(Question.is_active == True)
    if topic and topic != "Tümü":
        db_questions = db_questions.filter(Question.topic_id == topic)
    if difficulty:
        db_questions = db_questions.filter(Question.difficulty == difficulty)
    if question_type:
        db_questions = db_questions.filter(Question.question_type == _normalize_question_type(question_type))
    if bloom_level:
        db_questions = db_questions.filter(Question.bloom_level == bloom_level)
    if search:
        db_questions = db_questions.filter(Question.question_text.ilike(f"%{search}%"))

    questions_list = db_questions.all()
    if questions_list:
        return [
            StudentQuestion(
                id=q.question_id,
                topic=q.topic_id,
                question=q.question_text,
                options=q.options_json if q.options_json else [],
                question_type=q.question_type.value,
                difficulty=q.difficulty,
                bloom_level=q.bloom_level
            )
            for q in questions_list
        ]
        
    # Legacy fallback
    legacy = _load_full_questions()
    if topic and topic != "Tümü":
        legacy = [q for q in legacy if q["topic"] == topic]
        
    return [
        StudentQuestion(
            id=q["id"],
            topic=q["topic"],
            question=q["question"],
            options=q["options"]
        )
        for q in legacy
    ]


@router.post("/submit", response_model=SubmitResponse, status_code=status.HTTP_200_OK)
def submit_answers(body: SubmitRequest, current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)), db: Session = Depends(get_db)):
    db_questions = db.query(Question).filter(Question.question_id.in_(body.answers.keys())).all()
    
    # LEGACY PATH
    if not db_questions:
        question_map = {q["id"]: q for q in _load_full_questions()}
        results: List[QuestionFeedback] = []
        correct_count = 0

        for qid, selected in body.answers.items():
            if qid not in question_map: continue
            q = question_map[qid]
            is_correct = selected == q["correct_option"]
            if is_correct: correct_count += 1
            
            results.append(QuestionFeedback(
                id=q["id"],
                topic=q["topic"],
                question=q["question"],
                selected_option=selected,
                is_correct=is_correct,
                feedback=_safe_feedback(is_correct=is_correct)
            ))
            
        total = len(results)
        percentage = round((correct_count / total) * 100) if total > 0 else 0
        
        attempt = ExamResult(
            user_id=current_user.user_id,
            case_id="quiz_global",
            score=correct_count,
            max_score=total,
            details_json=json.dumps({"results": [r.model_dump() for r in results], "percentage": percentage}, ensure_ascii=False)
        )
        db.add(attempt)
        db.commit()
        db.refresh(attempt)
        
        return SubmitResponse(
            attempt_id=attempt.id,
            score=correct_count,
            total=total,
            percentage=percentage,
            overall_status="PUBLISHED",
            results=results
        )
        
    # S8B PATH
    q_map = {q.question_id: q for q in db_questions}
    attempt = QuizAttempt(user_id=current_user.user_id, total_score=0, max_score=0)
    db.add(attempt)
    db.flush()
    
    results = []
    total_score = 0
    total_max = 0
    has_pending = False
    
    for qid, selected in body.answers.items():
        if qid not in q_map: continue
        q = q_map[qid]
        
        total_max += q.max_score
        answer = QuizAnswer(
            attempt_id=attempt.id,
            question_id=q.id,
            student_response_text=selected
        )
        
        if q.question_type == QuestionType.MCQ:
            is_correct = selected == q.correct_option
            earned = q.max_score if is_correct else 0
            answer.auto_score = earned
            answer.grading_status = GradingStatus.PUBLISHED
            total_score += earned

            if not is_correct:
                _ensure_review_schedule(db, current_user.user_id, q)  # S10-C

            db.add(answer)
            db.flush()  # get answer.id for S10-B explanation link
            if not is_correct:
                _ensure_review_schedule(db, current_user.user_id, q)  # S10-C

            results.append(QuestionFeedback(
                id=q.question_id,
                topic=q.topic_id,
                question=q.question_text,
                question_type=q.question_type.value,
                selected_option=selected,
                is_correct=is_correct,
                feedback=_safe_feedback(is_correct=is_correct),
                grading_status=GradingStatus.PUBLISHED.value,
                answer_id=answer.id,  # S10-B
            ))
        else:
            answer.grading_status = GradingStatus.PENDING
            has_pending = True
            db.add(answer)

            results.append(QuestionFeedback(
                id=q.question_id,
                topic=q.topic_id,
                question=q.question_text,
                question_type=q.question_type.value,
                selected_option=selected,
                is_correct=None,
                feedback=None,
                grading_status=GradingStatus.PENDING.value
            ))
        
    if not has_pending:
        attempt.total_score = total_score
        attempt.max_score = total_max
        attempt.completed_at = datetime.utcnow()
        db.commit()
        db.refresh(attempt)
        perc = round((total_score / total_max) * 100) if total_max > 0 else 0
        return SubmitResponse(
            attempt_id=attempt.id,
            score=total_score,
            total=total_max,
            percentage=perc,
            overall_status="PUBLISHED",
            results=results
        )
    else:
        attempt.max_score = total_max
        db.commit()
        db.refresh(attempt)
        return SubmitResponse(
            attempt_id=attempt.id,
            score=None,
            total=total_max,
            percentage=None,
            overall_status="PENDING",
            results=results
        )


@router.get("/attempts/{attempt_id}", response_model=SubmitResponse, status_code=status.HTTP_200_OK)
def get_attempt(attempt_id: int, current_user: AuthenticatedUser = Depends(get_current_user_context), db: Session = Depends(get_db)):
    # Try S8B QuizAttempt first
    attempt = db.query(QuizAttempt).filter(QuizAttempt.id == attempt_id).first()
    if attempt:
        if current_user.role == UserRole.STUDENT and attempt.user_id != current_user.user_id:
            raise _forbidden()
        if current_user.role not in {UserRole.STUDENT, UserRole.INSTRUCTOR, UserRole.ADMIN}:
            raise _forbidden()
            
        results = []
        has_pending = False
        
        for ans in attempt.answers:
            q = ans.question
            if ans.grading_status != GradingStatus.PUBLISHED:
                has_pending = True
                
            feedback_item = QuestionFeedback(
                id=q.question_id,
                topic=q.topic_id,
                question=q.question_text,
                question_type=q.question_type.value,
                selected_option=ans.student_response_text,
                is_correct=None if q.question_type == QuestionType.OPEN_ENDED else (ans.auto_score > 0 if ans.auto_score is not None else False),
                feedback=_safe_feedback(is_correct=(ans.auto_score > 0 if ans.auto_score is not None else False)) if q.question_type == QuestionType.MCQ else None,
                grading_status=ans.grading_status.value
            )
            
            if ans.grading_status == GradingStatus.PUBLISHED:
                feedback_item.instructor_score = ans.instructor_score
                feedback_item.instructor_feedback = ans.instructor_feedback
                
            results.append(feedback_item)
            
        if has_pending:
            return SubmitResponse(
                attempt_id=attempt.id,
                score=None,
                total=attempt.max_score,
                percentage=None,
                overall_status="PENDING",
                results=results
            )
        else:
            perc = round((attempt.total_score / attempt.max_score) * 100) if attempt.max_score > 0 else 0
            return SubmitResponse(
                attempt_id=attempt.id,
                score=attempt.total_score,
                total=attempt.max_score,
                percentage=perc,
                overall_status="PUBLISHED",
                results=results
            )

    # Fallback to Legacy ExamResult
    legacy = db.query(ExamResult).filter(ExamResult.id == attempt_id).first()
    if not legacy or legacy.case_id != "quiz_global":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Quiz attempt not found")
        
    if current_user.role == UserRole.STUDENT and legacy.user_id != current_user.user_id:
        raise _forbidden()
    if current_user.role not in {UserRole.STUDENT, UserRole.INSTRUCTOR, UserRole.ADMIN}:
        raise _forbidden()
        
    payload = {}
    if legacy.details_json:
        try:
            payload = json.loads(legacy.details_json)
        except json.JSONDecodeError:
            pass
            
    raw_results = payload.get("results", []) if isinstance(payload, dict) else []
    safe_results = []
    if isinstance(raw_results, list):
        for item in raw_results:
            if isinstance(item, dict):
                try:
                    safe_results.append(
                        QuestionFeedback(
                            id=str(item.get("id", "")),
                            topic=str(item.get("topic", "")),
                            question=str(item.get("question", "")),
                            selected_option=item.get("selected_option"),
                            is_correct=bool(item.get("is_correct", False)),
                            feedback=str(item.get("feedback", "")),
                        )
                    )
                except Exception:
                    continue
                    
    total = int(legacy.max_score or 0)
    percentage = int(payload.get("percentage", round((legacy.score / total) * 100) if total > 0 else 0))
    
    return SubmitResponse(
        attempt_id=legacy.id,
        score=int(legacy.score or 0),
        total=total,
        percentage=percentage,
        overall_status="PUBLISHED",
        results=safe_results
    )


# ── Student Quiz History (T-5D) ──────────────────────────────────────────────

class AttemptListItem(BaseModel):
    attempt_id: int
    created_at: str
    total_score: int
    max_score: int
    percentage: Optional[int] = None
    question_count: int
    overall_status: str


@router.get("/my-attempts", response_model=List[AttemptListItem], status_code=status.HTTP_200_OK)
def list_my_attempts(
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)),
    db: Session = Depends(get_db),
):
    attempts = (
        db.query(QuizAttempt)
        .filter(QuizAttempt.user_id == current_user.user_id)
        .order_by(QuizAttempt.created_at.desc())
        .all()
    )
    items: List[AttemptListItem] = []
    for a in attempts:
        has_pending = any(
            ans.grading_status != GradingStatus.PUBLISHED for ans in a.answers
        )
        pct = round((a.total_score / a.max_score) * 100) if a.max_score > 0 and not has_pending else None
        items.append(AttemptListItem(
            attempt_id=a.id,
            created_at=a.created_at.isoformat() + "Z" if a.created_at else "",
            total_score=a.total_score,
            max_score=a.max_score,
            percentage=pct,
            question_count=len(a.answers),
            overall_status="PENDING" if has_pending else "PUBLISHED",
        ))
    return items


@router.get("/my-attempts/{attempt_id}", response_model=SubmitResponse, status_code=status.HTTP_200_OK)
def get_my_attempt_detail(
    attempt_id: int,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)),
    db: Session = Depends(get_db),
):
    attempt = db.query(QuizAttempt).filter(QuizAttempt.id == attempt_id).first()
    if not attempt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attempt not found")
    if attempt.user_id != current_user.user_id:
        raise _forbidden()

    results = []
    has_pending = False
    for ans in attempt.answers:
        q = ans.question
        if ans.grading_status != GradingStatus.PUBLISHED:
            has_pending = True
        fb = QuestionFeedback(
            id=q.question_id,
            topic=q.topic_id,
            question=q.question_text,
            question_type=q.question_type.value,
            selected_option=ans.student_response_text,
            is_correct=None if q.question_type == QuestionType.OPEN_ENDED else (ans.auto_score > 0 if ans.auto_score is not None else False),
            feedback=_safe_feedback(is_correct=(ans.auto_score > 0 if ans.auto_score is not None else False)) if q.question_type == QuestionType.MCQ else None,
            grading_status=ans.grading_status.value,
        )
        if ans.grading_status == GradingStatus.PUBLISHED:
            fb.instructor_score = ans.instructor_score
            fb.instructor_feedback = ans.instructor_feedback
        results.append(fb)

    if has_pending:
        return SubmitResponse(attempt_id=attempt.id, score=None, total=attempt.max_score, percentage=None, overall_status="PENDING", results=results)
    perc = round((attempt.total_score / attempt.max_score) * 100) if attempt.max_score > 0 else 0
    return SubmitResponse(attempt_id=attempt.id, score=attempt.total_score, total=attempt.max_score, percentage=perc, overall_status="PUBLISHED", results=results)


# ── S10-B: "Why this score?" Explainability ──────────────────────────────────

class RubricVersionSnapshot(BaseModel):
    version: int
    rubric_guide: str
    model_answer_outline: str
    created_at: Optional[str] = None


class AnswerExplanationResponse(BaseModel):
    answer_id: int
    question_id: str
    question_text: str
    question_type: str
    topic_id: str
    student_response: str
    auto_score: Optional[int] = None
    instructor_score: Optional[int] = None
    ai_score_suggestion: Optional[float] = None
    ai_score_rationale: Optional[str] = None
    max_score: int
    grading_status: str
    rubric_guide: Optional[str] = None
    rubric_version_snapshot: Optional[RubricVersionSnapshot] = None


@router.get(
    "/my-attempts/{attempt_id}/answers/{answer_id}/explanation",
    response_model=AnswerExplanationResponse,
    status_code=status.HTTP_200_OK,
    summary="Get 'Why this score?' explanation for a specific answer",
    description=(
        "Returns AI rationale, rubric guide, and rubric version snapshot for a "
        "specific quiz answer. Students can only access their own attempts. "
        "Returns 403 if the attempt belongs to another student. "
        "Returns 404 if attempt or answer not found."
    ),
)
def get_answer_explanation(
    attempt_id: int,
    answer_id: int,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT, UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
) -> AnswerExplanationResponse:
    attempt = db.query(QuizAttempt).filter(QuizAttempt.id == attempt_id).first()
    if not attempt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attempt not found")
    if current_user.role == UserRole.STUDENT and attempt.user_id != current_user.user_id:
        raise _forbidden()

    ans = db.query(QuizAnswer).filter(
        QuizAnswer.id == answer_id,
        QuizAnswer.attempt_id == attempt_id,
    ).first()
    if not ans:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Answer not found in this attempt")

    q = ans.question
    rubric_snapshot: Optional[RubricVersionSnapshot] = None
    if ans.rubric_version_id:
        rv = db.query(RubricVersion).filter(RubricVersion.id == ans.rubric_version_id).first()
        if rv:
            rubric_snapshot = RubricVersionSnapshot(
                version=rv.version,
                rubric_guide=rv.rubric_guide,
                model_answer_outline=rv.model_answer_outline,
                created_at=rv.created_at.isoformat() if rv.created_at else None,
            )

    return AnswerExplanationResponse(
        answer_id=ans.id,
        question_id=q.question_id,
        question_text=q.question_text,
        question_type=q.question_type.value,
        topic_id=q.topic_id,
        student_response=ans.student_response_text,
        auto_score=ans.auto_score,
        instructor_score=ans.instructor_score,
        ai_score_suggestion=ans.ai_score_suggestion,
        ai_score_rationale=ans.ai_score_rationale,
        max_score=q.max_score,
        grading_status=ans.grading_status.value,
        rubric_guide=q.rubric_guide,
        rubric_version_snapshot=rubric_snapshot,
    )


@router.get("/topics", status_code=status.HTTP_200_OK)
def get_topics(current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)), db: Session = Depends(get_db)):
    db_topics = [r[0] for r in db.query(Question.topic_id).filter(Question.is_active == True).distinct().all()]
    if db_topics:
        return ["Tümü"] + sorted(db_topics)

    questions = _load_full_questions()
    topics = sorted(set(q["topic"] for q in questions))
    return ["Tümü"] + topics


@router.get("/student/question-bank", response_model=List[QuestionBankEntry], status_code=status.HTTP_200_OK)
def get_student_question_bank(
    topic: Optional[str] = None,
    difficulty: Optional[str] = None,
    question_type: Optional[str] = None,
    bloom_level: Optional[str] = None,
    search: Optional[str] = None,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)),
    db: Session = Depends(get_db),
):
    q_query = db.query(Question).filter(Question.is_active == True)
    if topic and topic != "Tümü":
        q_query = q_query.filter(Question.topic_id == topic)
    if difficulty:
        q_query = q_query.filter(Question.difficulty == difficulty)
    if question_type:
        q_query = q_query.filter(Question.question_type == _normalize_question_type(question_type))
    if bloom_level:
        q_query = q_query.filter(Question.bloom_level == bloom_level)
    if search:
        q_query = q_query.filter(Question.question_text.ilike(f"%{search}%"))

    questions = q_query.all()

    # Fetch all answers this student has submitted in one query.
    user_answers = (
        db.query(QuizAnswer)
        .join(QuizAttempt)
        .filter(QuizAttempt.user_id == current_user.user_id)
        .all()
    )

    # Group by Question.id (integer PK) for O(1) lookup.
    answer_map: Dict[int, List] = defaultdict(list)
    for ans in user_answers:
        answer_map[ans.question_id].append(ans)

    result = []
    for q in questions:
        answers_for_q = answer_map.get(q.id, [])
        attempted = len(answers_for_q) > 0

        last_score = None
        if attempted:
            latest = max(answers_for_q, key=lambda a: a.id)
            raw = latest.instructor_score if latest.instructor_score is not None else latest.auto_score
            last_score = float(raw) if raw is not None else None

        result.append(QuestionBankEntry(
            question_id=q.question_id,
            question_text=q.question_text,
            question_type=q.question_type.value,
            topic_id=q.topic_id,
            bloom_level=q.bloom_level,
            difficulty=q.difficulty,
            max_score=q.max_score,
            options_json=q.options_json,
            attempt_summary=AttemptSummary(
                attempted=attempted,
                last_score=last_score,
                attempt_count=len(answers_for_q),
            ),
        ))

    return result


# ── Instructor Endpoints ──────────────────────────────────────────────────────

@router.get("/instructor/grading_queue", response_model=List[GradingQueueItem], status_code=status.HTTP_200_OK)
def get_grading_queue(current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)), db: Session = Depends(get_db)):
    pending_answers = db.query(QuizAnswer).join(Question).filter(
        QuizAnswer.grading_status != GradingStatus.PUBLISHED,
        Question.question_type == QuestionType.OPEN_ENDED
    ).all()
    
    return [
        GradingQueueItem(
            answer_id=ans.id,
            attempt_id=ans.attempt_id,
            question_id=ans.question.question_id,
            question_text=ans.question.question_text,
            student_response=ans.student_response_text,
            rubric_guide=ans.question.rubric_guide,
            model_answer_outline=ans.question.model_answer_outline,
            max_score=ans.question.max_score,
            submitted_at=ans.attempt.created_at.isoformat() if ans.attempt else None
        )
        for ans in pending_answers
    ]

@router.post("/instructor/grade/{answer_id}", status_code=status.HTTP_200_OK)
def submit_grade(answer_id: int, payload: GradeSubmission, current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)), db: Session = Depends(get_db)):
    ans = db.query(QuizAnswer).filter(QuizAnswer.id == answer_id).first()
    if not ans:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Answer not found")
        
    ans.instructor_score = payload.instructor_score
    ans.instructor_feedback = payload.instructor_feedback
    ans.graded_by_id = current_user.user_id
    ans.graded_at = datetime.utcnow()
    
    if payload.publish:
        ans.grading_status = GradingStatus.PUBLISHED
    else:
        ans.grading_status = GradingStatus.GRADED

    db.commit()

    # Create notification for the student when score is published
    if payload.publish:
        student_user_id = ans.attempt.user_id
        notification = Notification(
            user_id=student_user_id,
            type="score_published",
            payload_json={
                "answer_id": ans.id,
                "question_text": (ans.question.question_text or "")[:100],
                "score": payload.instructor_score,
                "max_score": ans.question.max_score,
                "graded_by": current_user.display_name,
            },
        )
        db.add(notification)
        db.commit()

    # Update Attempt Score if all answers are PUBLISHED
    attempt = ans.attempt
    if all(a.grading_status == GradingStatus.PUBLISHED for a in attempt.answers):
        total = sum((a.instructor_score or 0) + (a.auto_score or 0) for a in attempt.answers)
        attempt.total_score = total
        attempt.completed_at = datetime.utcnow()
        db.commit()

    return {"status": "success"}


# ── AI draft scoring endpoint (Sprint 4 — T-4A) ──────────────────────────────

class AIScoringResponse(BaseModel):
    answer_id: int
    suggested_score: float
    rationale: str
    scored_at: str
    max_score: int


@router.post(
    "/instructor/answers/{answer_id}/ai-score",
    response_model=AIScoringResponse,
    status_code=status.HTTP_200_OK,
    summary="Request AI draft score for an open-ended answer",
    description=(
        "Calls the LLM (HuggingFace Gemma) to generate a draft score for the "
        "given open-ended quiz answer. The draft is stored as ai_score_suggestion "
        "but never auto-promoted to instructor_score — the instructor must still "
        "accept or override it via the normal grading endpoint. "
        "Returns 404 if the answer does not exist, 400 if the question is MCQ, "
        "422 if the LLM call fails or returns an invalid response."
    ),
)
def request_ai_score_for_answer(
    answer_id: int,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    from app.services.oe_scoring_service import OEScoringError, request_ai_score

    ans = db.query(QuizAnswer).filter(QuizAnswer.id == answer_id).first()
    if not ans:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Answer not found")

    question = ans.question
    if question.question_type != QuestionType.OPEN_ENDED:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="AI scoring is only available for Open-Ended questions",
        )
    if not question.rubric_guide or not question.model_answer_outline:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Question is missing rubric_guide or model_answer_outline — cannot score",
        )

    import time as _time
    _t0 = _time.monotonic()
    try:
        result = request_ai_score(
            question_text=question.question_text,
            rubric_guide=question.rubric_guide,
            model_answer_outline=question.model_answer_outline,
            student_response=ans.student_response_text,
            max_score=question.max_score,
        )
    except OEScoringError as exc:
        latency = int((_time.monotonic() - _t0) * 1000)
        db.add(AIScoringLog(
            answer_id=answer_id,
            model_id="google/gemma-2-9b-it",
            status="error",
            error_message=str(exc),
            latency_ms=latency,
        ))
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"AI scoring failed: {exc}",
        )

    latency = int((_time.monotonic() - _t0) * 1000)
    ans.ai_score_suggestion = result.suggested_score
    ans.ai_score_rationale = result.rationale
    ans.ai_scored_at = datetime.utcnow()
    db.add(AIScoringLog(
        answer_id=answer_id,
        model_id="google/gemma-2-9b-it",
        status="success",
        latency_ms=latency,
        suggested_score=result.suggested_score,
    ))
    db.commit()

    return AIScoringResponse(
        answer_id=answer_id,
        suggested_score=result.suggested_score,
        rationale=result.rationale,
        scored_at=result.scored_at,
        max_score=question.max_score,
    )


@router.get(
    "/instructor/questions",
    response_model=List[InstructorQuestionSummary],
    status_code=status.HTTP_200_OK,
)
def get_instructor_open_ended_questions(
    topic: Optional[str] = None,
    question_type: Optional[str] = None,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    query = db.query(Question)
    if topic:
        query = query.filter(Question.topic_id == topic)
    if question_type:
        query = query.filter(Question.question_type == _normalize_question_type(question_type))

    questions = query.order_by(Question.created_at.desc(), Question.id.desc()).all()
    return [_question_to_summary(question) for question in questions]


@router.post(
    "/instructor/questions",
    response_model=InstructorQuestionSummary,
    status_code=status.HTTP_201_CREATED,
)
def create_instructor_open_ended_question(
    payload: InstructorQuestionCreateRequest,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    question_type = _normalize_question_type(payload.question_type)
    question_text = _normalize_text(payload.question_text)
    topic_id = _normalize_text(payload.topic_id)
    bloom_level = _normalize_text(payload.bloom_level)
    difficulty = _normalize_text(payload.difficulty)
    safety_category = _normalize_text(payload.safety_category)
    rubric_guide = _normalize_text(payload.rubric_guide)
    model_answer_outline = _normalize_text(payload.model_answer_outline)
    instructor_explanation = _normalize_text(payload.instructor_explanation)
    competency_areas = _normalize_str_list(payload.competency_areas)
    options = _normalize_str_list(payload.options)
    correct_option = _normalize_text(payload.correct_option)

    if not question_text:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Question text is required")
    if not topic_id:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Topic is required")
    if not bloom_level:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Bloom level is required")
    if not difficulty:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Difficulty is required")
    if not safety_category:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Safety category is required")
    if payload.max_score < 1 or payload.max_score > 100:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Max score must be between 1 and 100")

    if question_type == QuestionType.OPEN_ENDED:
        if not rubric_guide:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Rubric guide is required")
        if not model_answer_outline:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Model answer outline is required")
        if options:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Open-ended questions cannot include options")
        if correct_option:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Open-ended questions cannot include a correct option")
    else:
        if len(options) < 3:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="MCQ questions require at least 3 options")
        if len(options) != len(set(options)):
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="MCQ options must be unique")
        if not correct_option:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Correct option is required for MCQ questions")
        if correct_option not in options:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Correct option must match one of the provided options")

    requested_id = _normalize_text(payload.question_id)
    if requested_id and not re.fullmatch(r"[a-zA-Z0-9_-]+", requested_id):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Question ID may contain only letters, numbers, hyphens, and underscores",
        )
    question_id = requested_id or _generate_question_id(db, question_type, topic_id, question_text)
    if db.query(Question).filter(Question.question_id == question_id).first():
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Question ID already exists")

    question = Question(
        question_id=question_id,
        question_type=question_type,
        question_text=question_text,
        is_active=payload.is_active,
        topic_id=topic_id,
        competency_areas=competency_areas,
        bloom_level=bloom_level,
        difficulty=difficulty,
        safety_category=safety_category,
        unit_id=_normalize_text(payload.unit_id) or None,          # T-2A
        week_number=payload.week_number,                            # T-2A
        options_json=options or None,
        correct_option=correct_option or None,
        instructor_explanation=instructor_explanation or None,
        rubric_guide=rubric_guide or None,
        model_answer_outline=model_answer_outline or None,
        max_score=payload.max_score,
    )
    db.add(question)
    db.commit()
    db.refresh(question)

    return _question_to_summary(question)


# ── Composite score endpoint (Sprint 2 — T-2B) ───────────────────────────────

class ComponentScoreResponse(BaseModel):
    """API-safe representation of one score component."""
    available: bool
    earned: int
    max_possible: int
    pct: Optional[float]
    design_weight: float
    effective_weight: float


class CompositeScoreResponse(BaseModel):
    """
    Weighted composite score for the authenticated student.

    composite_pct is None when the student has no history across any component
    (true cold start). A value of 0.0 means attempts exist but zero points
    were earned.

    Weights when all components are available:
        MCQ          35 %
        Open-ended   40 %
        Case         25 %

    When a component is unavailable its design weight is redistributed
    proportionally across the remaining components.
    """
    mcq: ComponentScoreResponse
    open_ended: ComponentScoreResponse
    case: ComponentScoreResponse
    composite_pct: Optional[float]
    all_components_available: bool
    computed_at: str


def _result_to_response(result: CompositeScoreResult) -> CompositeScoreResponse:
    def _comp(c):
        return ComponentScoreResponse(
            available=c.available,
            earned=c.earned,
            max_possible=c.max_possible,
            pct=c.pct,
            design_weight=c.design_weight,
            effective_weight=c.effective_weight,
        )
    return CompositeScoreResponse(
        mcq=_comp(result.mcq),
        open_ended=_comp(result.open_ended),
        case=_comp(result.case),
        composite_pct=result.composite_pct,
        all_components_available=result.all_components_available,
        computed_at=result.computed_at,
    )


@router.get(
    "/my-score",
    response_model=CompositeScoreResponse,
    status_code=status.HTTP_200_OK,
    summary="Get my composite weighted score",
    description=(
        "Returns the student's weighted composite score across MCQ (35%), "
        "open-ended (40%), and case simulation (25%) components. "
        "Only published/graded records are counted. "
        "Missing components are flagged as unavailable — composite_pct is None "
        "only when the student has zero history across all components."
    ),
)
def get_my_composite_score(
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)),
    db: Session = Depends(get_db),
) -> CompositeScoreResponse:
    result = calculate_composite_score(current_user.user_id, db)
    return _result_to_response(result)


# ── Topic accuracy endpoint (Sprint 2 — T-2C) ────────────────────────────────

class TopicAccuracyItemResponse(BaseModel):
    """Per-topic MCQ accuracy for one topic."""
    topic_id: str
    topic_label: str
    earned: int
    max_possible: int
    pct: Optional[float]
    answered_count: int
    correct_count: int
    is_weak: bool


class TopicAccuracyResponse(BaseModel):
    """
    Per-topic MCQ accuracy breakdown for the authenticated student.

    Topics are sorted weakest-first (lowest pct first) so the most
    problematic areas appear at the top of the list.

    has_any_data is False when the student has no published MCQ answers
    yet (cold start). In that case, topics will be an empty list.
    """
    topics: List[TopicAccuracyItemResponse]
    has_any_data: bool
    computed_at: str


def _topic_result_to_response(result: TopicAccuracyResult) -> TopicAccuracyResponse:
    return TopicAccuracyResponse(
        topics=[
            TopicAccuracyItemResponse(
                topic_id=t.topic_id,
                topic_label=t.topic_label,
                earned=t.earned,
                max_possible=t.max_possible,
                pct=t.pct,
                answered_count=t.answered_count,
                correct_count=t.correct_count,
                is_weak=t.is_weak,
            )
            for t in result.topics
        ],
        has_any_data=result.has_any_data,
        computed_at=result.computed_at,
    )


@router.get(
    "/my-topic-accuracy",
    response_model=TopicAccuracyResponse,
    status_code=status.HTTP_200_OK,
    summary="Get my per-topic MCQ accuracy",
    description=(
        "Returns the student's MCQ accuracy broken down by topic. "
        "Only published answers are counted. "
        "Topics are sorted weakest-first (lowest accuracy first). "
        "A topic is marked weak when accuracy is below 60 %. "
        "has_any_data is False when the student has no MCQ history yet."
    ),
)
def get_my_topic_accuracy(
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)),
    db: Session = Depends(get_db),
) -> TopicAccuracyResponse:
    result = get_topic_accuracy(current_user.user_id, db)
    return _topic_result_to_response(result)


# ── Question-Case Mapping endpoint (Sprint 3 — T-3A) ─────────────────────────

class QuestionCaseMappingItem(BaseModel):
    """One theory-to-case link, enriched with question metadata."""
    id: int
    question_pk: int
    question_id: str
    question_type: str
    topic_id: str
    question_text: str
    case_id: str
    mapping_type: str
    review_status: str


class QuestionCaseMappingsResponse(BaseModel):
    """
    Theory-to-case mapping graph for the question bank.

    Each item links one Question to one CaseDefinition.
    The graph may be filtered by question_id, case_id, mapping_type,
    or review_status via query parameters.

    Accessible to instructors and admins only.
    """
    mappings: List[QuestionCaseMappingItem]
    total: int
    computed_at: str


def _mapping_result_to_response(result: MappingQueryResult) -> QuestionCaseMappingsResponse:
    return QuestionCaseMappingsResponse(
        mappings=[
            QuestionCaseMappingItem(
                id=m.id,
                question_pk=m.question_pk,
                question_id=m.question_id,
                question_type=m.question_type,
                topic_id=m.topic_id,
                question_text=m.question_text,
                case_id=m.case_id,
                mapping_type=m.mapping_type,
                review_status=m.review_status,
            )
            for m in result.mappings
        ],
        total=result.total,
        computed_at=result.computed_at,
    )


@router.get(
    "/question-case-mappings",
    response_model=QuestionCaseMappingsResponse,
    status_code=status.HTTP_200_OK,
    summary="List theory-to-case mappings",
    description=(
        "Returns all QuestionCaseMapping records joined with question metadata. "
        "Supports optional filtering by question_id, case_id, mapping_type, "
        "and review_status. Results are ordered by question_id then case_id. "
        "Accessible to instructors and admins only."
    ),
)
def get_question_case_mappings_endpoint(
    question_id: Optional[str] = None,
    case_id: Optional[str] = None,
    mapping_type: Optional[str] = None,
    review_status: Optional[str] = None,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
) -> QuestionCaseMappingsResponse:
    try:
        result = get_question_case_mappings(
            db,
            question_id=question_id,
            case_id=case_id,
            mapping_type=mapping_type,
            review_status=review_status,
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        )
    return _mapping_result_to_response(result)


# ── Write endpoints (T-3B) ────────────────────────────────────────────────────

class CreateMappingRequest(BaseModel):
    """Request body for creating a new theory-to-case mapping."""
    question_id: str = Field(
        ...,
        description="String identifier from Question.question_id (e.g. 'oral_path_001').",
    )
    case_id: str = Field(
        ...,
        description="Case identifier string to link (e.g. 'case_pericoronitis_01').",
    )
    mapping_type: str = Field(
        ...,
        description="Relationship type. One of: theory_support, case_reinforcement, assessment_link.",
    )
    review_status: Optional[str] = Field(
        default="unmapped",
        description="Review status. One of: approved, blocked_review_needed, unmapped. Defaults to 'unmapped'.",
    )


def _record_to_item(record) -> QuestionCaseMappingItem:
    return QuestionCaseMappingItem(
        id=record.id,
        question_pk=record.question_pk,
        question_id=record.question_id,
        question_type=record.question_type,
        topic_id=record.topic_id,
        question_text=record.question_text,
        case_id=record.case_id,
        mapping_type=record.mapping_type,
        review_status=record.review_status,
    )


@router.post(
    "/instructor/question-case-mappings",
    response_model=QuestionCaseMappingItem,
    status_code=status.HTTP_201_CREATED,
    summary="Create a theory-to-case mapping",
    description=(
        "Links a question to a clinical case. "
        "The question must already exist in the question bank. "
        "Returns 404 when question_id is not found. "
        "Returns 409 when the (question_id, case_id) pair already has a mapping. "
        "Returns 422 when mapping_type or review_status is not a valid enum value."
    ),
)
def create_question_case_mapping(
    payload: CreateMappingRequest,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
) -> QuestionCaseMappingItem:
    try:
        record = create_mapping(
            db,
            question_id=payload.question_id,
            case_id=payload.case_id,
            mapping_type=payload.mapping_type,
            review_status=payload.review_status or "unmapped",
        )
    except QuestionNotFoundError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    except DuplicateMappingError as exc:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc))
    return _record_to_item(record)


@router.delete(
    "/instructor/question-case-mappings/{mapping_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete a theory-to-case mapping",
    description=(
        "Permanently removes the QuestionCaseMapping row with the given id. "
        "Returns 404 when the mapping does not exist."
    ),
)
def delete_question_case_mapping(
    mapping_id: int,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    from app.services.question_case_mapping_service import (
        MappingNotFoundError,
        delete_mapping,
    )
    try:
        delete_mapping(db, mapping_id=mapping_id)
    except MappingNotFoundError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))


# ── Case Rubric endpoints (Sprint 3 — T-3C) ──────────────────────────────────

from app.services.case_rubric_service import (
    CaseNotFoundError,
    CaseRubric,
    DecisionPoint,
    get_all_case_rubrics,
    get_case_rubric,
    list_available_case_ids,
)


class DecisionPointResponse(BaseModel):
    target_action: str
    score: int
    rule_outcome: str
    is_critical: bool
    safety_category: Optional[str]
    competency_tags: List[str]
    rubric_level: str


class CaseRubricResponse(BaseModel):
    case_id: str
    total_max_score: int
    critical_count: int
    positive_count: int
    penalty_count: int
    computed_at: str
    decision_points: List[DecisionPointResponse]


def _rubric_to_response(rubric: CaseRubric) -> CaseRubricResponse:
    return CaseRubricResponse(
        case_id=rubric.case_id,
        total_max_score=rubric.total_max_score,
        critical_count=rubric.critical_count,
        positive_count=rubric.positive_count,
        penalty_count=rubric.penalty_count,
        computed_at=rubric.computed_at,
        decision_points=[
            DecisionPointResponse(
                target_action=dp.target_action,
                score=dp.score,
                rule_outcome=dp.rule_outcome,
                is_critical=dp.is_critical,
                safety_category=dp.safety_category,
                competency_tags=dp.competency_tags,
                rubric_level=dp.rubric_level,
            )
            for dp in rubric.decision_points
        ],
    )


@router.get(
    "/case-rubrics",
    response_model=List[CaseRubricResponse],
    summary="List rubrics for all clinical cases",
)
def list_case_rubrics(
    current_user: AuthenticatedUser = Depends(require_roles(
        UserRole.STUDENT, UserRole.INSTRUCTOR, UserRole.ADMIN
    )),
):
    return [_rubric_to_response(r) for r in get_all_case_rubrics()]


@router.get(
    "/case-rubrics/{case_id}",
    response_model=CaseRubricResponse,
    summary="Get rubric for a single clinical case",
)
def get_single_case_rubric(
    case_id: str,
    current_user: AuthenticatedUser = Depends(require_roles(
        UserRole.STUDENT, UserRole.INSTRUCTOR, UserRole.ADMIN
    )),
):
    try:
        rubric = get_case_rubric(case_id)
    except CaseNotFoundError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    return _rubric_to_response(rubric)


@router.get(
    "/case-rubrics-index",
    response_model=List[str],
    summary="List all case IDs that have scoring rules",
)
def list_case_rubric_ids(
    current_user: AuthenticatedUser = Depends(require_roles(
        UserRole.STUDENT, UserRole.INSTRUCTOR, UserRole.ADMIN
    )),
):
    return list_available_case_ids()


# =============================================================================
# T-4B: Rubric Versioning Endpoints
# =============================================================================

from app.services.rubric_version_service import (
    snapshot_rubric,
    get_rubric_versions,
    get_rubric_version,
    RubricVersionError,
    RubricVersionInfo,
)


class RubricSnapshotRequest(BaseModel):
    """Payload for publishing a new rubric version."""
    rubric_guide: str = Field(..., min_length=1, description="Updated rubric guide text.")
    model_answer_outline: str = Field(..., min_length=1, description="Updated model answer outline.")
    change_notes: Optional[str] = Field(None, description="Optional description of what changed.")


class RubricVersionResponse(BaseModel):
    """API-facing representation of a RubricVersion snapshot."""
    id: int
    question_id: int
    version: int
    rubric_guide: str
    model_answer_outline: str
    change_notes: Optional[str]
    created_by: str
    created_at: str


def _rv_to_response(info: RubricVersionInfo) -> RubricVersionResponse:
    return RubricVersionResponse(
        id=info.id,
        question_id=info.question_id,
        version=info.version,
        rubric_guide=info.rubric_guide,
        model_answer_outline=info.model_answer_outline,
        change_notes=info.change_notes,
        created_by=info.created_by,
        created_at=info.created_at,
    )


@router.post(
    "/instructor/questions/{question_id}/rubric-snapshot",
    response_model=RubricVersionResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Publish a new rubric version snapshot for a question",
)
def publish_rubric_snapshot(
    question_id: int,
    payload: RubricSnapshotRequest,
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
):
    """
    Saves an immutable snapshot of the rubric for audit purposes and
    updates the question's live rubric_guide and model_answer_outline.
    """
    try:
        info = snapshot_rubric(
            db,
            question_id=question_id,
            rubric_guide=payload.rubric_guide,
            model_answer_outline=payload.model_answer_outline,
            change_notes=payload.change_notes,
            created_by=current_user.email,
        )
    except RubricVersionError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    return _rv_to_response(info)


@router.get(
    "/instructor/questions/{question_id}/rubric-versions",
    response_model=List[RubricVersionResponse],
    summary="List all rubric version snapshots for a question (newest first)",
)
def list_rubric_versions(
    question_id: int,
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
):
    try:
        versions = get_rubric_versions(db, question_id=question_id)
    except RubricVersionError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    return [_rv_to_response(v) for v in versions]


@router.get(
    "/instructor/rubric-versions/{version_id}",
    response_model=RubricVersionResponse,
    summary="Get a single rubric version snapshot by ID",
)
def get_single_rubric_version(
    version_id: int,
    db: Session = Depends(get_db),
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
):
    try:
        info = get_rubric_version(db, version_id=version_id)
    except RubricVersionError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    return _rv_to_response(info)


# ── Bulk Question Actions (T-5C) ────────────────────────────────────────────

class BulkQuestionActionRequest(BaseModel):
    question_ids: List[int]
    action: str = Field(..., pattern="^(archive|activate|update_unit|update_week)$")
    value: Optional[str] = None


class BulkActionResult(BaseModel):
    affected: int
    action: str


@router.patch(
    "/instructor/questions/bulk",
    response_model=BulkActionResult,
    status_code=status.HTTP_200_OK,
)
def bulk_update_questions(
    body: BulkQuestionActionRequest,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    questions = db.query(Question).filter(Question.id.in_(body.question_ids)).all()
    if not questions:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No questions found")

    for q in questions:
        if body.action == "archive":
            q.is_archived = True
            q.is_active = False
        elif body.action == "activate":
            q.is_archived = False
            q.is_active = True
        elif body.action == "update_unit":
            q.unit_id = body.value
        elif body.action == "update_week":
            q.week_number = int(body.value) if body.value else None

    db.commit()
    return BulkActionResult(affected=len(questions), action=body.action)


class ImportPreviewItem(BaseModel):
    question_id: str
    question_type: str
    question_text: str
    topic_id: str
    difficulty: str
    status: str


class ImportResult(BaseModel):
    added: int
    updated: int
    skipped: int
    errors: List[str]
    preview: List[ImportPreviewItem]


@router.post(
    "/instructor/import",
    response_model=ImportResult,
    status_code=status.HTTP_200_OK,
)
async def import_questions_endpoint(
    file: UploadFile,
    upsert: bool = False,
    dry_run: bool = False,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    import csv
    import io

    content = await file.read()
    text = content.decode("utf-8")

    questions: List[dict] = []
    filename = file.filename or ""
    if filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            questions.append(dict(row))
    else:
        try:
            parsed = json.loads(text)
            if not isinstance(parsed, list):
                return ImportResult(added=0, updated=0, skipped=0, errors=["JSON must be a top-level array"], preview=[])
            questions = parsed
        except json.JSONDecodeError as e:
            return ImportResult(added=0, updated=0, skipped=0, errors=[f"Invalid JSON: {e}"], preview=[])

    from scripts.import_questions import validate_question, REQUIRED_FIELDS

    errors: List[str] = []
    preview: List[ImportPreviewItem] = []
    added = 0
    updated = 0
    skipped = 0

    for q in questions:
        qid = str(q.get("question_id", "")).strip()
        errs = validate_question(q)
        if errs:
            for e in errs:
                errors.append(f"{qid or '<missing>'}:  {e}")
            preview.append(ImportPreviewItem(
                question_id=qid or "<missing>",
                question_type=str(q.get("question_type", "")),
                question_text=str(q.get("question_text", ""))[:80],
                topic_id=str(q.get("topic_id", "")),
                difficulty=str(q.get("difficulty", "")),
                status="error",
            ))
            continue

        existing = db.query(Question).filter(Question.question_id == qid).first()
        if existing and not upsert:
            skipped += 1
            preview.append(ImportPreviewItem(
                question_id=qid, question_type=q["question_type"],
                question_text=q["question_text"][:80], topic_id=q["topic_id"],
                difficulty=q["difficulty"], status="skipped",
            ))
            continue

        action = "updated" if existing else "added"
        if existing:
            updated += 1
        else:
            added += 1

        if not dry_run:
            from scripts.import_questions import _apply_question_fields
            if existing:
                _apply_question_fields(existing, q)
            else:
                model = Question(question_id=qid)
                _apply_question_fields(model, q)
                db.add(model)

        preview.append(ImportPreviewItem(
            question_id=qid, question_type=q["question_type"],
            question_text=q["question_text"][:80], topic_id=q["topic_id"],
            difficulty=q["difficulty"], status=action,
        ))

    if not dry_run and not errors:
        db.commit()

    return ImportResult(added=added, updated=updated, skipped=skipped, errors=errors, preview=preview)


@router.get(
    "/instructor/questions/export",
    status_code=status.HTTP_200_OK,
)
def export_questions_csv(
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    import csv
    import io
    from starlette.responses import StreamingResponse

    questions = db.query(Question).filter(Question.is_active == True).all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["question_id", "question_type", "topic_id", "difficulty", "bloom_level", "question_text", "max_score", "unit_id", "week_number"])
    for q in questions:
        writer.writerow([q.question_id, q.question_type.value, q.topic_id, q.difficulty, q.bloom_level, q.question_text, q.max_score, q.unit_id or "", q.week_number or ""])

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=questions_export.csv"},
    )


# ── T-7B: Instructor Grade Report Export ─────────────────────────────────────

@router.get("/instructor/grade-report", status_code=status.HTTP_200_OK)
def export_grade_report(
    format: str = "csv",
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    import csv
    import io
    from starlette.responses import StreamingResponse

    students = db.query(User).filter(User.role == UserRole.STUDENT, User.is_archived == False).all()

    topic_ids = sorted(set(
        t[0] for t in db.query(Question.topic_id).distinct().all()
    ))

    rows = []
    for student in students:
        attempts = db.query(QuizAttempt).filter(QuizAttempt.user_id == student.user_id).all()
        attempt_ids = [a.id for a in attempts]
        if not attempt_ids:
            continue

        answers = (
            db.query(QuizAnswer)
            .filter(QuizAnswer.attempt_id.in_(attempt_ids), QuizAnswer.grading_status == GradingStatus.PUBLISHED)
            .all()
        )
        if not answers:
            continue

        topic_correct: dict = defaultdict(int)
        topic_total: dict = defaultdict(int)
        ai_scores = []
        total_earned = 0
        total_max = 0

        for ans in answers:
            q = ans.question
            if not q:
                continue
            tid = q.topic_id
            if q.question_type == QuestionType.MCQ:
                topic_total[tid] = topic_total.get(tid, 0) + 1
                if ans.auto_score and ans.auto_score > 0:
                    topic_correct[tid] = topic_correct.get(tid, 0) + 1
            score = ans.instructor_score if ans.instructor_score is not None else (ans.auto_score or 0)
            total_earned += score
            total_max += q.max_score
            if ans.ai_score_suggestion is not None:
                ai_scores.append(ans.ai_score_suggestion)

        row = {
            "user_id": student.user_id,
            "display_name": student.display_name,
        }
        for tid in topic_ids:
            tot = topic_total.get(tid, 0)
            cor = topic_correct.get(tid, 0)
            row[f"topic_{tid}_pct"] = round(cor / tot * 100, 1) if tot > 0 else ""

        row["ai_score_avg"] = round(sum(ai_scores) / len(ai_scores), 2) if ai_scores else ""
        row["total_earned"] = total_earned
        row["total_max"] = total_max
        row["final_pct"] = round(total_earned / total_max * 100, 1) if total_max > 0 else 0
        rows.append(row)

    if format == "xlsx":
        try:
            import openpyxl
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Grade Report"
            if rows:
                headers = list(rows[0].keys())
                ws.append(headers)
                for r in rows:
                    ws.append([r.get(h, "") for h in headers])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=grade_report.xlsx"},
            )
        except ImportError:
            pass

    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=grade_report.csv"},
    )


# ── T-7C: Student Personal PDF Report ────────────────────────────────────────

@router.get("/my-report", status_code=status.HTTP_200_OK)
def get_my_report(
    format: str = "pdf",
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)),
    db: Session = Depends(get_db),
):
    import io
    from starlette.responses import StreamingResponse
    from fpdf import FPDF

    attempts = db.query(QuizAttempt).filter(QuizAttempt.user_id == current_user.user_id).all()
    attempt_ids = [a.id for a in attempts]

    answers = (
        db.query(QuizAnswer)
        .filter(QuizAnswer.attempt_id.in_(attempt_ids), QuizAnswer.grading_status == GradingStatus.PUBLISHED)
        .all()
    ) if attempt_ids else []

    topic_correct: dict = defaultdict(int)
    topic_total: dict = defaultdict(int)
    total_earned = 0
    total_max = 0

    for ans in answers:
        q = ans.question
        if not q:
            continue
        tid = q.topic_id
        if q.question_type == QuestionType.MCQ:
            topic_total[tid] = topic_total.get(tid, 0) + 1
            if ans.auto_score and ans.auto_score > 0:
                topic_correct[tid] = topic_correct.get(tid, 0) + 1
        score = ans.instructor_score if ans.instructor_score is not None else (ans.auto_score or 0)
        total_earned += score
        total_max += q.max_score

    final_pct = round(total_earned / total_max * 100, 1) if total_max > 0 else 0

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "DentAI - Kisisel Performans Raporu", ln=True, align="C")
    pdf.ln(5)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Ogrenci: {current_user.display_name} ({current_user.user_id})", ln=True)
    pdf.cell(0, 8, f"Tarih: {datetime.utcnow().strftime('%Y-%m-%d')}", ln=True)
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 10, "Genel Ozet", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Toplam Puan: {total_earned} / {total_max}  ({final_pct}%)", ln=True)
    pdf.cell(0, 8, f"Yanitlanan Soru Sayisi: {len(answers)}", ln=True)
    pdf.ln(5)

    if topic_total:
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 10, "Konu Bazli Dogruluk (MCQ)", ln=True)
        pdf.set_font("Helvetica", "", 10)
        for tid in sorted(topic_total.keys()):
            tot = topic_total[tid]
            cor = topic_correct.get(tid, 0)
            pct = round(cor / tot * 100, 1) if tot > 0 else 0
            label = TOPIC_MAP.get(tid, tid)
            bar_len = int(pct / 5)
            bar = "#" * bar_len + "." * (20 - bar_len)
            pdf.cell(0, 7, f"  {label}: {cor}/{tot} ({pct}%)  [{bar}]", ln=True)
        pdf.ln(3)

    weak_topics = [
        tid for tid, tot in topic_total.items()
        if tot > 0 and (topic_correct.get(tid, 0) / tot) < 0.5
    ]
    strong_topics = [
        tid for tid, tot in topic_total.items()
        if tot > 0 and (topic_correct.get(tid, 0) / tot) >= 0.8
    ]

    if strong_topics:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 9, "Guclu Alanlar:", ln=True)
        pdf.set_font("Helvetica", "", 10)
        for tid in strong_topics:
            pdf.cell(0, 7, f"  + {TOPIC_MAP.get(tid, tid)}", ln=True)
        pdf.ln(2)

    if weak_topics:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 9, "Gelistirilmesi Gereken Alanlar:", ln=True)
        pdf.set_font("Helvetica", "", 10)
        for tid in weak_topics:
            pdf.cell(0, 7, f"  - {TOPIC_MAP.get(tid, tid)}", ln=True)

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=rapor_{current_user.user_id}.pdf"},
    )


# ── T-7D: Question Stats Dashboard ───────────────────────────────────────────

class QuestionStatItem(BaseModel):
    question_id: int
    question_text_short: str
    topic_id: str
    difficulty: str
    total_answers: int
    correct_count: int
    correct_pct: float
    avg_ai_score: Optional[float] = None
    avg_instructor_score: Optional[float] = None
    ai_human_delta: Optional[float] = None


@router.get(
    "/instructor/question-stats",
    response_model=List[QuestionStatItem],
    status_code=status.HTTP_200_OK,
)
def get_question_stats(
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    questions = db.query(Question).filter(Question.is_active == True).all()
    result = []
    for q in questions:
        answers = db.query(QuizAnswer).filter(QuizAnswer.question_id == q.id).all()
        if not answers:
            continue
        total = len(answers)
        correct = 0
        ai_scores = []
        instructor_scores = []
        for ans in answers:
            if q.question_type == QuestionType.MCQ and ans.auto_score and ans.auto_score > 0:
                correct += 1
            if ans.ai_score_suggestion is not None:
                ai_scores.append(ans.ai_score_suggestion)
            if ans.instructor_score is not None:
                instructor_scores.append(float(ans.instructor_score))

        avg_ai = round(sum(ai_scores) / len(ai_scores), 2) if ai_scores else None
        avg_ins = round(sum(instructor_scores) / len(instructor_scores), 2) if instructor_scores else None
        delta = round(abs(avg_ai - avg_ins), 2) if avg_ai is not None and avg_ins is not None else None

        result.append(QuestionStatItem(
            question_id=q.id,
            question_text_short=(q.question_text or "")[:80],
            topic_id=q.topic_id,
            difficulty=q.difficulty,
            total_answers=total,
            correct_count=correct,
            correct_pct=round(correct / total * 100, 1) if total > 0 else 0,
            avg_ai_score=avg_ai,
            avg_instructor_score=avg_ins,
            ai_human_delta=delta,
        ))

    result.sort(key=lambda x: x.correct_pct)
    return result


@router.get(
    "/instructor/ai-vs-human-delta",
    response_model=List[QuestionStatItem],
    status_code=status.HTTP_200_OK,
)
def get_ai_vs_human_delta(
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.INSTRUCTOR, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    all_stats = get_question_stats(current_user=current_user, db=db)
    with_delta = [s for s in all_stats if s.ai_human_delta is not None]
    with_delta.sort(key=lambda x: x.ai_human_delta or 0, reverse=True)
    return with_delta[:10]


# ── S10-C: Spaced Repetition Scheduler ───────────────────────────────────────

class ReviewScheduleItem(BaseModel):
    id: int
    question_id: str
    question_text: str
    topic_id: str
    due_date: str
    interval_days: int
    ease_factor: float
    repetitions: int
    last_reviewed_at: Optional[str] = None


class SubmitReviewRequest(BaseModel):
    rating: int = Field(..., ge=0, le=5, description="SM-2 rating 0 (fail) – 5 (easy)")


class SubmitReviewResponse(BaseModel):
    id: int
    next_due_date: str
    next_interval_days: int
    repetitions: int


@router.get(
    "/my-review-schedule",
    response_model=List[ReviewScheduleItem],
    status_code=status.HTTP_200_OK,
    summary="Get due review items for the current student (S10-C)",
    description=(
        "Returns questions due for spaced repetition review today or earlier. "
        "Items are ordered by due_date ascending (most overdue first). "
        "When the student has no scheduled items yet, an empty list is returned. "
        "New review entries are created automatically when a student answers a "
        "question incorrectly via the /quiz/submit endpoint."
    ),
)
def get_my_review_schedule(
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)),
    db: Session = Depends(get_db),
) -> List[ReviewScheduleItem]:
    now = datetime.utcnow()
    items = (
        db.query(ReviewSchedule)
        .filter(
            ReviewSchedule.user_id == current_user.user_id,
            ReviewSchedule.due_date <= now,
        )
        .order_by(ReviewSchedule.due_date.asc())
        .all()
    )
    result = []
    for item in items:
        q = item.question
        if not q or not q.is_active:
            continue
        result.append(ReviewScheduleItem(
            id=item.id,
            question_id=q.question_id,
            question_text=q.question_text,
            topic_id=q.topic_id,
            due_date=item.due_date.isoformat() + "Z",
            interval_days=item.interval_days,
            ease_factor=item.ease_factor,
            repetitions=item.repetitions,
            last_reviewed_at=item.last_reviewed_at.isoformat() + "Z" if item.last_reviewed_at else None,
        ))
    return result


@router.post(
    "/my-review-schedule/{item_id}/result",
    response_model=SubmitReviewResponse,
    status_code=status.HTTP_200_OK,
    summary="Submit a review result and advance the SM-2 schedule (S10-C)",
)
def submit_review_result(
    item_id: int,
    body: SubmitReviewRequest,
    current_user: AuthenticatedUser = Depends(require_roles(UserRole.STUDENT)),
    db: Session = Depends(get_db),
) -> SubmitReviewResponse:
    from app.services.spaced_repetition import next_review_state

    item = db.query(ReviewSchedule).filter(
        ReviewSchedule.id == item_id,
        ReviewSchedule.user_id == current_user.user_id,
    ).first()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Review item not found")

    now = datetime.utcnow()
    new_state = next_review_state(
        repetitions=item.repetitions,
        interval_days=item.interval_days,
        ease_factor=item.ease_factor,
        rating=body.rating,
        reviewed_at=now,
    )
    item.repetitions = new_state.repetitions
    item.interval_days = new_state.interval_days
    item.ease_factor = new_state.ease_factor
    item.due_date = new_state.due_date
    item.last_reviewed_at = now
    db.commit()

    return SubmitReviewResponse(
        id=item.id,
        next_due_date=new_state.due_date.isoformat() + "Z",
        next_interval_days=new_state.interval_days,
        repetitions=new_state.repetitions,
    )
